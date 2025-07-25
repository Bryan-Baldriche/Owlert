# Subnet shenanigans
discovered_subnet = None

COMMON_PORTS = [
     21, 22, 23, 25, 53, 67, 68, 69, 80, 110, 111, 123, 135, 137, 138, 139,
     143, 161, 162, 179, 389, 443, 445, 465, 514, 515, 993, 995, 1080, 1194,
     1433, 1434, 1521, 1723, 2049, 2121, 3306, 3389, 3690, 4444, 5060, 5432,
     5900, 5985, 5986, 6379, 8080, 8443, 8888, 9000, 9090, 9200, 27017
]

#Imported Goods (tarif free)

import socket
import threading
import psutil
import ipaddress
import os
import platform
from OwlertV3.utilities import user_input, clear_screen
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import openpyxl
import sys
import concurrent.futures
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import ssl
import csv

import re

_illegal = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def clean_banner(text: str) -> str:
    return _illegal.sub(".", text) if text else ""


print_lock = threading.Lock()


def main():
    discovered_subnet = None

    while True:
        print("\n📡 Owlert Network Recon 📡\n")
        print("1. Host Subnet Discovery")
        print("2. Specific Port Host Discovery")
        print("3. Scan All Ports for Hosts (Incredibly Slow!)")
        print("4. Fast Subnet-Wide Scan (Common Ports Scan)")
        print("5. Service Enumeration")
        print("6. exit")
        print("\n type 'help' at anytime for a list of additional commands")


        choice = user_input("📡 Network Recon: ")

        if choice == "1":
            host_subnet_discovery_func()
            subnet_scan()
        elif choice == "2":
            subnet_scan()
            host_finder()
        elif choice == "3":
            subnet_scan()
            all_port_scan_func()
        elif choice == "4":
            subnet_scan()
            quick_scanner()
        elif choice == "5":
            service_enum_func()
        elif choice == "exit" or choice == "back to home" or choice == "6":
            break
        else:
            print("Invalid option, please try again")

# Determines Host OS
def host_subnet_discovery_func():
    operating_system = platform.system()
    if operating_system == "Linux":
        print(f"\nHost Discovery - Running on 🐧{operating_system}\n")
    elif operating_system == "Windows":
        print(f"\nHost Discovery - Running on 🪟{operating_system}\n")
    elif operating_system == "Darwin":
        print(f"\nHost Discovery - Running on 🍎{operating_system}\n")
    else:
        print("\nunable to identify operating system!")

# Finds out host subnet
def subnet_scan():
    global discovered_subnet

    interfaces = psutil.net_if_addrs()

    for iface_name, addrs in interfaces.items():
        for addr in addrs:
            if addr.family == socket.AF_INET:
                ip = addr.address
                netmask = addr.netmask

                if ip.startswith("127.") or netmask is None:
                    continue

                try:
                    subnet = ipaddress.IPv4Network(f"{ip}/{netmask}", strict=False)
                    print(f"Detected Host: {iface_name} ({ip}/{netmask})")
                    discovered_subnet = subnet
                    return
                except Exception as e:
                    print(f"Skipped {iface_name} due to error: {e}")
                    continue

    print("No valid network interface found! cannot determine subnet, please check all connections!")
    discovered_subnet = None

# Scans a specified port for hosts on the subnet
def host_finder(timeout: float = 0.5):
    global discovered_subnet

    port_in = user_input("Enter port to scan for hosts [default 80]: ").strip()
    try:
        port = int(port_in) if port_in else 80
    except ValueError:
        print("Invalid input; defaulting to port 80.")
        port = 80

    # ── ask network scope ────────────────────────────────────────────────
    scope = user_input("Scan scope /8, /16, or /24 [default /16]: ").strip()
    if scope not in {"8", "16", "24", "/8", "/16", "/24"}:
        scope = "16"

    octets = discovered_subnet.network_address.exploded.split(".")
    if scope == "8" or scope == "/8":
        cidr = f"{octets[0]}.0.0.0/8"
    elif scope == "16" or scope == "/16":
        cidr = f"{octets[0]}.{octets[1]}.0.0/16"
    elif scope == "24" or scope == "/24":
        cidr = f"{octets[0]}.{octets[1]}.{octets[2]}.0/24"
    else:  # /24
        cidr = f"{octets[0]}.{octets[1]}.{octets[2]}.0/24"

    target_net = ipaddress.IPv4Network(cidr, strict=False)
    hosts = [str(h) for h in target_net.hosts()]
    total = len(hosts)

    print(f"\nScanning {target_net} on port {port}  (256 threads)…")
    print("Press Ctrl+C to stop early.\n")

    socket.setdefaulttimeout(timeout)
    start_time = time.time()
    live_hosts, counter = [], 0
    lock = threading.Lock()

    def probe(ip: str):
        nonlocal counter
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex((ip, port)) == 0:
                    with lock:
                        live_hosts.append(ip)
                        print(f"\n[📡] Host Found! | {ip}:{port}", flush=True)
        except KeyboardInterrupt:
            pass
        finally:
            with lock:
                counter += 1
                elapsed = time.time() - start_time
                print(f"[{counter:>6}/{total}] {elapsed:6.1f}s  Scanning {ip}:{port}",
                      end="\r", flush=True)

    try:
        with ThreadPoolExecutor(max_workers=256) as pool:
            pool.map(probe, hosts)
    except KeyboardInterrupt:
        print("\nScan interrupted by user!")

    # ── summary ──────────────────────────────────────────────────────────
    if live_hosts:
        print(f"\nDiscovery complete: {len(live_hosts)} host(s) detected.")
    else:
        print("\nNo live hosts detected.")

    # ── export to Excel ──────────────────────────────────────────────────
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Port-Specific Scan"
        ws.append(["IP Address", f"Port {port} open"])
        for ip in sorted(live_hosts):
            ws.append([ip, "Yes"])
        wb.save("port_specific_scan_results.xlsx")
        print("Results saved to port_specific_scan_results.xlsx")
    except Exception as e:
        print(f"Could not write Excel report: {e}")

    return live_hosts


def host_finder_result():
    print("\nFinal Host Discovery Result:\n")
    hosts = host_finder()
    if hosts:
        print("\nLive Hosts:")
        for h in hosts:
            print(f" • {h}")
    else:
        print("No hosts found.")

def all_port_scan_func():
    global discovered_subnet

    print("Starting full scan of all ports across all hosts on network...\n")

    scope = user_input("Scan scope?  /8, /16, or /24 [default /16]: ").strip()
    if scope not in {"8", "16", "24", "/8", "/16", "/24"}:
        scope = "16"

    base_octets = discovered_subnet.network_address.exploded.split(".")
    if scope == "8" or scope == "/8":
        cidr_str = f"{base_octets[0]}.0.0.0/8"
    elif scope == "16" or scope == "/16":
        cidr_str = f"{base_octets[0]}.{base_octets[1]}.0.0/16"
    elif scope == "24" or scope == "/24":  # "/24"
        cidr_str = f"{base_octets[0]}.{base_octets[1]}.{base_octets[2]}.0/24"
    else:  # "/24"
        cidr_str = f"{base_octets[0]}.{base_octets[1]}.{base_octets[2]}.0/24"

    target_net = ipaddress.IPv4Network(cidr_str, strict=False)
    hosts = [str(h) for h in target_net.hosts()]

    print(f"\n🚀 Deep scan of ALL ports for network {target_net} "
          f"({len(hosts)} hosts)\n")

    start = time.time()
    results = []

    # ── per‑port worker -------------------------------------------------------
    def scan_port(ip, port, open_port_list):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.2)
                if s.connect_ex((ip, port)) == 0:
                    open_port_list.append(port)
        except Exception:
            pass

    # ── per‑host worker -------------------------------------------------------
    def scan_host(ip: str):
        open_ports = []
        with print_lock:
            print(f"⎈ Scanning {ip} …", flush=True)

        with ThreadPoolExecutor(max_workers=100) as port_pool:
            port_pool.map(lambda p: scan_port(ip, p, open_ports), range(1, 65536))

        if open_ports:
            with print_lock:
                print(f"[+] {ip} open: {open_ports}")
            return ip, sorted(open_ports)
        return None

    # ── scan all hosts --------------------------------------------------------
    try:
        with ThreadPoolExecutor(max_workers=256) as host_pool:
            for ip_ports in host_pool.map(scan_host, hosts):
                if ip_ports:
                    results.append(ip_ports)
    except KeyboardInterrupt:
        print("\nScan interrupted by user!")
        return

    elapsed = time.time() - start
    print(f"\nDeep scan finished in {elapsed / 60:.2f} minutes.")
    print(f"Hosts with open ports: {len(results)}")

    # ── export Excel ----------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "All‑Port Results"
    ws.append(["IP Address", "Open Ports"])
    for ip, ports in sorted(results, key=lambda x: x[0]):
        ws.append([ip, ", ".join(map(str, ports))])
    out_path = "all_port_scan_results.xlsx"
    wb.save(out_path)
    print(f"Results saved to {out_path}")

def quick_scanner():
    global discovered_subnet

    scope = user_input("Scan scope?  /8, /16, or /24 [default /16]: ").strip()
    if scope not in {"8", "16", "24", "/8", "/16", "/24"}:
        scope = "16"

    # derive target network from current interface IP
    octets = discovered_subnet.network_address.exploded.split(".")
    if scope == "8" or scope == "/8":
        cidr = f"{octets[0]}.0.0.0/8"
    elif scope == "16" or scope == "/16":
        cidr = f"{octets[0]}.{octets[1]}.0.0/16"
    elif scope == "24" or scope == "/24":
        cidr = f"{octets[0]}.{octets[1]}.{octets[2]}.0/24"
    else:  # /24
        cidr = f"{octets[0]}.{octets[1]}.{octets[2]}.0/24"

    target_network = ipaddress.IPv4Network(cidr, strict=False)
    print(f"\nStarting Quick + Conditional Deep Scan over {target_network}")
    print("Quick phase: common ports. Deep phase: full 1‑65535 if host responds.\n")

    start_time = time.time()
    results = []
    responsive_hosts = 0
    deep_scanned_hosts = 0
    aborted = False

    # ── per‑IP scanning routine ────────────────────────────────────────────
    def scan_ip(ip: str):
        open_ports = []
        found_common = False

        with print_lock:
            print(f"Performing Quick Scan of {ip}", flush=True)

        # Quick scan of common ports
        for port in COMMON_PORTS:
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(0.25)
                    if s.connect_ex((ip, port)) == 0:
                        open_ports.append(port)
                        found_common = True
                        with print_lock:
                            elapsed = time.time() - start_time
                            print(f"Host found! | {ip}:{port} "
                                  f"| Now performing Deep Scan… (t+{elapsed:.1f}s)",
                                  flush=True)
                        break
            except Exception:
                continue

        if not found_common:
            with print_lock:
                print(f"No Host Found on {ip}", flush=True)
            return ip, open_ports, False  # skip deep scan

        # Deep scan (all remaining ports) with port threads
        def check_port(p):
            if p in COMMON_PORTS:
                return
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.settimeout(0.15)
                    if s.connect_ex((ip, p)) == 0:
                        with print_lock:
                            print(f"[+] {ip}:{p} open", flush=True)
                        open_ports.append(p)
            except Exception:
                pass

        with ThreadPoolExecutor(max_workers=256) as port_pool:
            port_pool.map(check_port, range(1, 65536))

        with print_lock:
            print(f"[+] {ip} -> Open ports: {sorted(open_ports)}", flush=True)

        return ip, sorted(open_ports), True

    hosts = [str(h) for h in target_network.hosts()]

    try:
        with ThreadPoolExecutor(max_workers=256) as executor:
            future_map = {executor.submit(scan_ip, ip): ip for ip in hosts}

            for future in as_completed(future_map):
                ip, ports, deep_used = future.result()
                if ports:
                    responsive_hosts += 1
                    if deep_used:
                        deep_scanned_hosts += 1
                    results.append((ip, ports))

    except KeyboardInterrupt:
        aborted = True
        with print_lock:
            print("\nScan interrupted by user. Finalizing partial results…")

    duration = time.time() - start_time

    # ── summary ────────────────────────────────────────────────────────────
    print("\n========== Quick Scanner Summary ==========")
    print(f"Target network:       {target_network}")
    print(f"Duration (seconds):   {duration:.2f}")
    print(f"Responsive hosts:     {responsive_hosts}")
    print(f"Deep‑scanned hosts:   {deep_scanned_hosts}")
    print(f"Total hosts reported: {len(results)}")
    print(f"Status:               {'ABORTED' if aborted else 'COMPLETE'}")
    print("===========================================")

    # ── export to Excel ────────────────────────────────────────────────────
    out_name = "quick_scan_results_partial.xlsx" if aborted else "quick_scan_results.xlsx"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Quick Scan Results"
        ws.append(["IP Address", "Open Ports"])
        for ip, ports in sorted(results, key=lambda x: x[0]):
            ws.append([ip, ", ".join(map(str, ports)) if ports else "None"])
        wb.save(out_name)
        print(f"\nResults saved to: {out_name}")
    except Exception as e:
        print(f"Could not write Excel report: {e}")

    return results

def banner_reader(sock, length=128):
    try:
        return sock.recv(length).decode(errors="ignore").strip()
    except Exception:
        return ""

def enumerator(ip, port, timeout=1):
    try:
        with socket.create_connection((ip, port), timeout=timeout) as s:
            # send a small probe if bannerless
            s.sendall(b"HEAD / HTTP/1.0\r\n\r\n")
            banner = banner_reader(s)
            if not banner:
                banner = "No banner / timeout"
    except Exception:
        banner = "No banner / timeout"
    return port, banner

def service_enum_func():
    """
    Service Enumeration:
      • User can enter an IP (and ports) manually,
        OR press Enter to load default XLSX (port_scan_results.xlsx),
        OR type a custom .xlsx filename.
      • XLSX must have columns: IP Address | Open Ports (comma separated).
      • Results saved to service_enum_results.xlsx & .csv
    """
    print("\n🛰️  Service Enumeration")

    choice = user_input(
        "Target IP, or press [Enter] to load 'quick_scan_results.xlsx',\n"
        "or type another .xlsx file path to load: "
    ).strip()

    targets = {}  # {ip: [int ports]}

    # Case 1: user entered what looks like an IP
    if choice and all(part.isdigit() for part in choice.split(".")) and len(choice.split(".")) == 4:
        ip = choice
        port_list = user_input("Comma-separated ports (e.g. 22,80,443): ").strip()
        try:
            ports = [int(p.strip()) for p in port_list.split(",") if p.strip().isdigit()]
        except ValueError:
            print("❌ Invalid port list. Aborting.")
            return
        targets[ip] = ports

    else:
        # Use default or custom file
        xlsx_path = choice or "quick_scan_results.xlsx"
        if not os.path.isfile(xlsx_path):
            print(f"❌ File '{xlsx_path}' not found.")
            return
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            # Expect header row: IP Address | Open Ports
            for row in ws.iter_rows(min_row=2, values_only=True):
                ip, ports_str = row
                if not ip or not ports_str or ports_str == "None":
                    continue
                try:
                    ports = [int(p.strip()) for p in str(ports_str).split(",") if p.strip().isdigit()]
                    targets[str(ip)] = ports
                except ValueError:
                    continue
        except Exception as e:
            print("❌ Could not read Excel file:", e)
            return

    # Enumerate
    print(f"\nEnumerating services on {len(targets)} host(s)…\n")
    enum_results = []  # (ip, port, banner)

    for ip, ports in targets.items():
        print(f"→ {ip} ({len(ports)} port(s))")
        for port in ports:
            p, banner = enumerator(ip, port)
            print(f"  {p:<5} : {banner[:100]}")
            enum_results.append((ip, p, banner))

    # Save results
    if enum_results:
        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Service Enum"
        ws.append(["IP Address", "Port", "Banner / Info"])
        for ip, p, banner in enum_results:
            ws.append([ip, p, clean_banner(banner)])
        wb.save("service_enum_results.xlsx")

        # CSV
        with open("service_enum_results.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["IP Address", "Port", "Banner / Info"])
            writer.writerows(enum_results)

        print("\n✅ Results saved to service_enum_results.xlsx and service_enum_results.csv")
    else:
        print("No banners found / nothing to save.")


if __name__ == "__main__":
    main()